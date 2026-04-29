from __future__ import annotations

import hashlib
import hmac
import pathlib
import secrets
import time
import uuid

from fastapi import FastAPI, HTTPException, Response, Request, Cookie
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse

# Import routers from each service, mounting them under distinct prefixes.
from services.ingestion.app.api import router as ingestion_router
from services.normalizer.app.api import router as normalizer_router
from services.mapping.app.api import router as mapping_router
from services.scoring.app.api import router as scoring_router
from services.reporting.app.api import router as reporting_router

# ---------- new defriends services (consent, behavioral, remediation, AI) --
# Each is import-guarded so a missing optional dep never breaks the core
# 5-stage pipeline. Failures are logged to stderr but don't abort startup.
try:
    from services.consent.app.api import router as consent_router
    _HAS_CONSENT = True
except Exception as _e:                                   # pragma: no cover
    import sys; print(f"[defriends] consent router unavailable: {_e}", file=sys.stderr)
    _HAS_CONSENT = False
try:
    from services.behavioral.app.api import router as behavioral_router
    _HAS_BEHAVIORAL = True
except Exception as _e:                                   # pragma: no cover
    import sys; print(f"[defriends] behavioral router unavailable: {_e}", file=sys.stderr)
    _HAS_BEHAVIORAL = False
try:
    from services.remediation.app.api import router as remediation_router
    _HAS_REMEDIATION = True
except Exception as _e:                                   # pragma: no cover
    import sys; print(f"[defriends] remediation router unavailable: {_e}", file=sys.stderr)
    _HAS_REMEDIATION = False
try:
    from services.ai_assistant.app.api import router as ai_app_router
    _HAS_AI_APP = True
except Exception as _e:                                   # pragma: no cover
    import sys; print(f"[defriends] ai_assistant router unavailable: {_e}", file=sys.stderr)
    _HAS_AI_APP = False
try:
    from services.security_core.middleware import install as _install_security_middleware
    _HAS_SECURITY_MW = True
except Exception as _e:                                   # pragma: no cover
    import sys; print(f"[defriends] security middleware unavailable: {_e}", file=sys.stderr)
    _HAS_SECURITY_MW = False

_STATIC_DIR = pathlib.Path(__file__).resolve().parent / "static"

app = FastAPI(
    title="defriends",
    description="defriends - authorized device security assessment workspace.",
    version="1.0.0",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

# Install the security middleware FIRST so every downstream router inherits
# rate-limiting, security headers, body-size caps, and secret redaction.
if _HAS_SECURITY_MW:
    _install_security_middleware(app)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://www.dirtybots.com",
        "http://localhost:8080",
        "http://127.0.0.1:8080",
        "http://localhost:3000",
        "http://127.0.0.1:8000",
        "http://localhost:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount all service routers (each already has /v1 prefix)
app.include_router(ingestion_router, tags=["ingestion"])
app.include_router(normalizer_router, tags=["normalizer"])
app.include_router(mapping_router, tags=["mapping"])
app.include_router(scoring_router, tags=["scoring"])
app.include_router(reporting_router, tags=["reporting"])

# ---------- new defriends services ----------------------------------------
if _HAS_CONSENT:
    app.include_router(consent_router)
if _HAS_BEHAVIORAL:
    app.include_router(behavioral_router)
if _HAS_REMEDIATION:
    app.include_router(remediation_router)
if _HAS_AI_APP:
    app.include_router(ai_app_router)


def _require_page_session(session_token: str | None, next_url: str):
    session = _get_session(session_token)
    if session:
        return None
    return RedirectResponse(url=f"/login?next={next_url}", status_code=303)


@app.get("/", tags=["gateway"])
def root():
    return RedirectResponse(url="/login")


@app.get("/login", response_class=HTMLResponse, tags=["gateway"])
def login_page():
    html_path = _STATIC_DIR / "login.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@app.get("/ui", response_class=HTMLResponse, tags=["gateway"])
def dashboard_ui(session_token: str | None = Cookie(default=None)):
    guard = _require_page_session(session_token, "/ui")
    if guard:
        return guard
    html_path = _STATIC_DIR / "dashboard.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@app.get("/user", response_class=HTMLResponse, tags=["gateway"])
def user_portal(session_token: str | None = Cookie(default=None)):
    guard = _require_page_session(session_token, "/user")
    if guard:
        return guard
    html_path = _STATIC_DIR / "user_dashboard.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@app.get("/docs", response_class=HTMLResponse, tags=["gateway"])
def api_docs(session_token: str | None = Cookie(default=None)):
    guard = _require_page_session(session_token, "/docs")
    if guard:
        return guard
    docs_path = _STATIC_DIR / "docs.html"
    return HTMLResponse(content=docs_path.read_text(encoding="utf-8"))


@app.get("/api", tags=["gateway"])
def api_info(session_token: str | None = Cookie(default=None)):
    if not _get_session(session_token):
        return {
            "service": "defriends API",
            "status": "login_required",
            "login": "/login",
            "docs": "/login?next=/docs",
            "health": "/health",
        }
    return {
        "service": "defriends API",
        "version": "0.1.0",
        "ui": "/ui",
        "docs": "/docs",
        "health": "/health",
        "endpoints": {
            "consent": "/v1/consent",
            "behavioral": "/v1/behavioral",
            "remediation": "/v1/remediation",
            "ai_app": "/v1/ai/app",
        },
        "website": "https://www.dirtybots.com",
    }


@app.get("/health", tags=["gateway"])
def health():
    return {
        "status": "ok",
        "service": "defriends-unified",
        "services": {
            "ingestion": "ok",
            "normalizer": "ok",
            "mapping": "ok",
            "scoring": "ok",
            "reporting": "ok",
            "consent": "ok" if _HAS_CONSENT else "unavailable",
            "behavioral": "ok" if _HAS_BEHAVIORAL else "unavailable",
            "remediation": "ok" if _HAS_REMEDIATION else "unavailable",
            "ai_assistant": "ok" if _HAS_AI_APP else "unavailable",
        },
        "security": {
            "middleware": "enabled" if _HAS_SECURITY_MW else "disabled",
        },
    }


@app.get("/v1/masterpiece/manifest", tags=["gateway"])
def masterpiece_manifest(session_token: str | None = Cookie(default=None)):
    """Return the product ideology and live capability map for the dashboard."""
    if not _get_session(session_token):
        raise HTTPException(status_code=401, detail="Authentication required")
    pillars = [
        {
            "name": "Consent is the first control",
            "signal": "Every collection path starts from a revocable receipt.",
            "score": 100 if _HAS_CONSENT else 70,
        },
        {
            "name": "Short memory by default",
            "signal": "Seven-day retention is the baseline; extensions require fresh permission.",
            "score": 96,
        },
        {
            "name": "Evidence becomes attacker language",
            "signal": "CVE, CWE, exposure, and behavior are mapped to MITRE ATT&CK techniques.",
            "score": 98,
        },
        {
            "name": "Risk is explained before it is ranked",
            "signal": "CVSS, EPSS, KEV, reachability, and internet exposure are visible inputs.",
            "score": 95,
        },
        {
            "name": "Fixes are reversible promises",
            "signal": "Remediation defaults to dry-run and keeps rollback steps close to the finding.",
            "score": 100 if _HAS_REMEDIATION else 75,
        },
        {
            "name": "Plain language is a security feature",
            "signal": "The local assistant explains findings without sending prompts to an outside model.",
            "score": 100 if _HAS_AI_APP else 72,
        },
    ]
    live_services = {
        "consent": _HAS_CONSENT,
        "behavioral": _HAS_BEHAVIORAL,
        "remediation": _HAS_REMEDIATION,
        "ai_assistant": _HAS_AI_APP,
        "security_middleware": _HAS_SECURITY_MW,
    }
    live_count = sum(1 for enabled in live_services.values() if enabled)
    pillar_score = round(sum(p["score"] for p in pillars) / len(pillars))
    readiness_score = round((pillar_score * 0.75) + ((live_count / len(live_services)) * 100 * 0.25))
    return {
        "name": "defriends Trust Manifest",
        "ideology": "Authenticated workspace for reviewing assessment evidence and practical next steps.",
        "uniqueness_note": (
            "The workspace focuses on authorized assessment, evidence-backed results, "
            "plain-language guidance, and accountable next actions."
        ),
        "readiness_score": readiness_score,
        "pillars": pillars,
        "live_services": live_services,
        "signature": "authorized review active",
    }


# ---------------------------------------------------------------------------
# Authentication & User Management (in-memory)
# ---------------------------------------------------------------------------
_SECRET_KEY = secrets.token_hex(32)
_users: dict[str, dict] = {
    # Pre-seeded admin account
    "admin@dirtybots.com": {
        "email": "admin@dirtybots.com",
        "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
        "role": "admin",
        "name": "Admin",
        "mfa_secret": None,
        "mfa_enabled": False,
        "phone": "",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    },
    "user@dirtybots.com": {
        "email": "user@dirtybots.com",
        "password_hash": hashlib.sha256("user123".encode()).hexdigest(),
        "role": "user",
        "name": "User",
        "mfa_secret": None,
        "mfa_enabled": False,
        "phone": "",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    },
}
_sessions: dict[str, dict] = {}  # token -> {email, role, created_at}
_mfa_pending: dict[str, dict] = {}  # temp_token -> {email, code, expires_at}


def _hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()


def _create_session(email: str, role: str) -> str:
    token = secrets.token_urlsafe(48)
    _sessions[token] = {
        "email": email,
        "role": role,
        "created_at": time.time(),
    }
    return token


def _get_session(token: str | None) -> dict | None:
    if not token:
        return None
    session = _sessions.get(token)
    if session and (time.time() - session["created_at"]) < 86400:  # 24h expiry
        return session
    if session:
        del _sessions[token]
    return None


def _generate_totp_code() -> str:
    """Generate a 6-digit TOTP-like code for MFA."""
    import random
    return f"{random.randint(100000, 999999)}"


@app.post("/v1/auth/register", tags=["auth"])
def register(body: dict):
    """Register a new user account."""
    email = (body.get("email") or "").strip().lower()
    password = body.get("password", "")
    name = body.get("name", "")
    phone = body.get("phone", "")
    role = body.get("role", "user")

    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Valid email is required")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if email in _users:
        raise HTTPException(status_code=409, detail="Email already registered")
    if role not in ("user", "admin"):
        role = "user"

    _users[email] = {
        "email": email,
        "password_hash": _hash_pw(password),
        "role": role,
        "name": name,
        "mfa_secret": None,
        "mfa_enabled": False,
        "phone": phone,
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }
    return {"registered": True, "email": email, "role": role}


@app.post("/v1/auth/login", tags=["auth"])
def login(body: dict):
    """Authenticate and get a session token. Returns MFA challenge if enabled."""
    email = (body.get("email") or "").strip().lower()
    password = body.get("password", "")

    user = _users.get(email)
    if not user or user["password_hash"] != _hash_pw(password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    if user.get("mfa_enabled"):
        code = _generate_totp_code()
        temp_token = secrets.token_urlsafe(32)
        _mfa_pending[temp_token] = {
            "email": email,
            "code": code,
            "expires_at": time.time() + 300,  # 5 min
        }
        # In production, send via SMS/email/authenticator
        return {
            "mfa_required": True,
            "mfa_token": temp_token,
            "mfa_hint": f"Code sent to {user.get('phone', 'your device')}",
            "_dev_code": code,  # DEV ONLY — remove in production
        }

    token = _create_session(email, user["role"])
    return {
        "authenticated": True,
        "token": token,
        "role": user["role"],
        "name": user["name"],
        "email": email,
    }


@app.post("/v1/auth/mfa/verify", tags=["auth"])
def verify_mfa(body: dict):
    """Verify MFA code and complete login."""
    mfa_token = body.get("mfa_token", "")
    code = body.get("code", "")

    pending = _mfa_pending.get(mfa_token)
    if not pending:
        raise HTTPException(status_code=400, detail="Invalid or expired MFA token")
    if time.time() > pending["expires_at"]:
        del _mfa_pending[mfa_token]
        raise HTTPException(status_code=400, detail="MFA code expired")
    if pending["code"] != code:
        raise HTTPException(status_code=401, detail="Invalid MFA code")

    del _mfa_pending[mfa_token]
    email = pending["email"]
    user = _users[email]
    token = _create_session(email, user["role"])
    return {
        "authenticated": True,
        "token": token,
        "role": user["role"],
        "name": user["name"],
        "email": email,
    }


@app.post("/v1/auth/mfa/enable", tags=["auth"])
def enable_mfa(body: dict):
    """Enable MFA for a user. Accepts phone number or authenticator setup."""
    token = body.get("token", "")
    phone = body.get("phone", "")
    session = _get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = _users.get(session["email"])
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user["mfa_enabled"] = True
    user["phone"] = phone
    # In production, generate TOTP secret for authenticator apps
    totp_secret = secrets.token_hex(20)
    user["mfa_secret"] = totp_secret
    return {
        "mfa_enabled": True,
        "method": "totp",
        "totp_secret": totp_secret,
        "message": "MFA enabled. Use an authenticator app or SMS verification.",
    }


@app.post("/v1/auth/logout", tags=["auth"])
def logout(body: dict):
    """Invalidate a session token."""
    token = body.get("token", "")
    if token in _sessions:
        del _sessions[token]
    return {"logged_out": True}


@app.get("/v1/auth/session", tags=["auth"])
def check_session(token: str = ""):
    """Check if a session token is valid."""
    session = _get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    user = _users.get(session["email"], {})
    return {
        "valid": True,
        "email": session["email"],
        "role": session["role"],
        "name": user.get("name", ""),
        "mfa_enabled": user.get("mfa_enabled", False),
    }


@app.get("/v1/auth/users", tags=["auth"])
def list_users(token: str = ""):
    """List all users (admin only)."""
    session = _get_session(token)
    if not session or session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return {
        "users": [
            {k: v for k, v in u.items() if k != "password_hash" and k != "mfa_secret"}
            for u in _users.values()
        ]
    }


# ---------------------------------------------------------------------------
# In-memory Agent Registry — tracks deployed agents
# ---------------------------------------------------------------------------
_agent_registry: dict[str, dict] = {}  # agent_id -> {meta}


@app.post("/v1/agents/register", tags=["agents"])
def register_agent(body: dict):
    """Register a deployed agent (called by the agent or dashboard)."""
    agent_id = body.get("agent_id") or f"agt-{uuid.uuid4().hex[:12]}"
    _agent_registry[agent_id] = {
        "agent_id": agent_id,
        "asset_id": body.get("asset_id", "unknown"),
        "org_id": body.get("org_id", "unknown"),
        "platform": body.get("platform", "unknown"),
        "os_name": body.get("os_name", ""),
        "os_version": body.get("os_version", ""),
        "environment": body.get("environment", "prod"),
        "deploy_mode": body.get("deploy_mode", "local"),
        "collectors": body.get("collectors", []),
        "status": "active",
        "registered_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "last_heartbeat": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "scan_count": 0,
    }
    return {"registered": True, "agent_id": agent_id}


@app.post("/v1/agents/{agent_id}/heartbeat", tags=["agents"])
def agent_heartbeat(agent_id: str, body: dict = {}):
    """Update agent heartbeat."""
    agent = _agent_registry.get(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    agent["last_heartbeat"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    agent["status"] = body.get("status", "active")
    return {"ok": True}


@app.get("/v1/agents", tags=["agents"])
def list_agents():
    """List all registered agents."""
    return {
        "agents": list(
            sorted(
                _agent_registry.values(),
                key=lambda a: a.get("registered_at", ""),
                reverse=True,
            )
        )
    }


@app.get("/v1/agents/export/excel", tags=["agents"])
def export_agents_excel():
    """Export all deployed agents as an Excel (.xlsx) file."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Deployed Agents"

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C6AF0", end_color="7C6AF0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Agent ID", "Asset ID", "Organization", "Platform", "OS Name",
        "OS Version", "Environment", "Deploy Mode", "Status",
        "Collectors", "Scan Count", "Registered At", "Last Heartbeat",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    agents = sorted(
        _agent_registry.values(),
        key=lambda a: a.get("registered_at", ""),
        reverse=True,
    )
    for agent in agents:
        ws.append([
            agent.get("agent_id", ""),
            agent.get("asset_id", ""),
            agent.get("org_id", ""),
            agent.get("platform", ""),
            agent.get("os_name", ""),
            agent.get("os_version", ""),
            agent.get("environment", ""),
            agent.get("deploy_mode", ""),
            agent.get("status", ""),
            ", ".join(agent.get("collectors", [])),
            agent.get("scan_count", 0),
            agent.get("registered_at", ""),
            agent.get("last_heartbeat", ""),
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.border = thin_border
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="dirtybots_deployed_agents.xlsx"'},
    )


# ---------------------------------------------------------------------------
# In-memory report store – keeps generated PDF reports for timeline download
# ---------------------------------------------------------------------------
_report_store: dict[str, dict] = {}  # report_id -> {meta, pdf_bytes}


@app.post("/v1/reports", tags=["reports"])
def store_report(body: dict):
    """Generate a PDF and store it in the report timeline."""
    import io
    try:
        from services.reporting.app.pdf_renderer import render_report_pdf
    except ModuleNotFoundError:
        raise HTTPException(status_code=503, detail="PDF renderer unavailable")

    report_id = body.get("report_id") or f"rpt-{uuid.uuid4().hex[:12]}"
    buf = io.BytesIO()
    try:
        render_report_pdf(body, buf)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    asset = body.get("asset", {})
    summary = body.get("summary", {})
    scan_context = body.get("scan_context", {})
    provenance = body.get("provenance", {})
    _report_store[report_id] = {
        "report_id": report_id,
        "generated_at": body.get("generated_at", time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())),
        "asset_id": asset.get("asset_id", "unknown"),
        "org_id": asset.get("org_id", "unknown"),
        "environment": asset.get("environment", "unknown"),
        "platform": scan_context.get("platform", "unknown"),
        "os_name": scan_context.get("os_name", provenance.get("os", "")),
        "os_version": scan_context.get("os_version", ""),
        "scan_mode": scan_context.get("scan_mode", "manual"),
        "deploy_mode": scan_context.get("deploy_mode", "local"),
        "risk_score": summary.get("overall_risk_score", 0),
        "findings_total": summary.get("findings_total", 0),
        "pdf_bytes": buf.getvalue(),
    }

    # Auto-register or update the agent
    agent_key = f"{asset.get('org_id', '')}-{asset.get('asset_id', '')}-{scan_context.get('platform', '')}"
    if agent_key not in _agent_registry:
        _agent_registry[agent_key] = {
            "agent_id": agent_key,
            "asset_id": asset.get("asset_id", "unknown"),
            "org_id": asset.get("org_id", "unknown"),
            "platform": scan_context.get("platform", "unknown"),
            "os_name": scan_context.get("os_name", provenance.get("os", "")),
            "os_version": scan_context.get("os_version", ""),
            "environment": asset.get("environment", "prod"),
            "deploy_mode": scan_context.get("deploy_mode", "local"),
            "collectors": scan_context.get("collectors", []),
            "status": "active",
            "registered_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "last_heartbeat": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "scan_count": 1,
        }
    else:
        _agent_registry[agent_key]["last_heartbeat"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        _agent_registry[agent_key]["scan_count"] = _agent_registry[agent_key].get("scan_count", 0) + 1
        _agent_registry[agent_key]["status"] = "active"

    return {
        "stored": True,
        "report_id": report_id,
        "download_url": f"/v1/reports/{report_id}/pdf",
    }


@app.get("/v1/reports", tags=["reports"])
def list_reports():
    """Return metadata for all stored reports (without PDF bytes)."""
    items = []
    for rid, entry in sorted(_report_store.items(), key=lambda x: x[1].get("generated_at", ""), reverse=True):
        items.append({k: v for k, v in entry.items() if k != "pdf_bytes"})
    return {"reports": items}


@app.get("/v1/reports/{report_id}/pdf", tags=["reports"])
def download_report(report_id: str):
    """Download a previously stored PDF report."""
    entry = _report_store.get(report_id)
    if not entry:
        raise HTTPException(status_code=404, detail=f"Report {report_id} not found")
    return Response(
        content=entry["pdf_bytes"],
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="dirtybots_{report_id}.pdf"'},
    )


# ---------------------------------------------------------------------------
# AI Assistant — domain-aware security intelligence engine
# ---------------------------------------------------------------------------

_AI_KNOWLEDGE = {
    "cve_db": {
        "CVE-2024-3094": {"title": "XZ Utils Backdoor", "severity": "critical", "cvss": 10.0, "remediation": "Upgrade xz-utils to >= 5.6.2. Rebuild all container images. This is a supply-chain backdoor actively exploited in the wild (KEV-listed).", "mitre": "T1195.002 Supply Chain Compromise"},
        "CVE-2023-44487": {"title": "HTTP/2 Rapid Reset", "severity": "high", "cvss": 7.5, "remediation": "Upgrade HTTP/2 server (nginx >= 1.25.3, Go >= 1.21.3). Apply rate limiting on HTTP/2 streams.", "mitre": "T1499.001 Endpoint DoS"},
        "CVE-2023-45853": {"title": "zlib Integer Overflow", "severity": "critical", "cvss": 9.8, "remediation": "Upgrade zlib to >= 1.3.1. Affects any application using MiniZip.", "mitre": "T1203 Exploitation for Client Execution"},
        "CVE-2021-44228": {"title": "Log4Shell", "severity": "critical", "cvss": 10.0, "remediation": "Upgrade Log4j to >= 2.17.1. Set -Dlog4j2.formatMsgNoLookups=true as interim mitigation.", "mitre": "T1190 Exploit Public-Facing Application"},
    },
    "mitre_techniques": {
        # Initial Access
        "T1190": {"name": "Exploit Public-Facing Application", "tactic": "Initial Access", "description": "Adversaries exploit vulnerabilities in internet-facing systems (web servers, databases, firewalls) to gain initial access. Examples: SQL injection, RCE in web apps, exploiting unpatched services.", "mitigations": ["Application Isolation and Sandboxing", "Exploit Protection", "Network Segmentation", "Privileged Account Management", "Update Software", "Vulnerability Scanning"], "platforms": ["Linux", "Windows", "macOS", "Containers"]},
        "T1195": {"name": "Supply Chain Compromise", "tactic": "Initial Access", "description": "Adversaries manipulate products or delivery mechanisms prior to receipt by the end consumer. This includes compromise of software (T1195.002), hardware, or dependencies.", "mitigations": ["Audit", "Update Software", "Vulnerability Scanning"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1195.001": {"name": "Supply Chain Compromise: Compromise Software Dependencies", "tactic": "Initial Access", "description": "Adversaries compromise software dependencies (npm, PyPI, Maven packages) to inject malicious code that executes when developers install or update packages.", "mitigations": ["Audit", "Update Software", "Vulnerability Scanning", "SBOM Tracking"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1195.002": {"name": "Supply Chain Compromise: Compromise Software Supply Chain", "tactic": "Initial Access", "description": "Adversaries compromise a software vendor's build/delivery pipeline to inject malicious code into legitimate software updates. Example: XZ Utils backdoor (CVE-2024-3094), SolarWinds.", "mitigations": ["Code Signing Verification", "SBOM Tracking", "Dependency Pinning", "Build Provenance Attestation (SLSA)"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1078": {"name": "Valid Accounts", "tactic": "Initial Access", "description": "Adversaries use stolen or compromised credentials to gain initial access, bypass access controls, or escalate privileges. Sub-techniques: Default Accounts (.001), Domain Accounts (.002), Local Accounts (.003), Cloud Accounts (.004).", "mitigations": ["MFA", "Password Policies", "Privileged Account Management", "User Account Management"], "platforms": ["Linux", "Windows", "macOS", "Azure AD", "Google Workspace"]},
        "T1566": {"name": "Phishing", "tactic": "Initial Access", "description": "Adversaries send messages with malicious attachments or links to gain access. Sub-techniques: Spearphishing Attachment (.001), Link (.002), via Service (.003).", "mitigations": ["Antivirus/Antimalware", "Network Intrusion Prevention", "Software Configuration", "User Training"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1133": {"name": "External Remote Services", "tactic": "Initial Access", "description": "Adversaries leverage external-facing remote services (VPN, Citrix, RDP) to gain access using valid credentials.", "mitigations": ["Disable or Remove Feature", "Limit Access to Resource Over Network", "MFA", "Network Segmentation"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1199": {"name": "Trusted Relationship", "tactic": "Initial Access", "description": "Adversaries exploit trusted third-party relationships (MSPs, supply chain partners) to gain access to target networks.", "mitigations": ["Network Segmentation", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        # Execution
        "T1059": {"name": "Command and Scripting Interpreter", "tactic": "Execution", "description": "Adversaries use command-line interfaces and scripting interpreters for execution. Sub-techniques: PowerShell (.001), AppleScript (.002), Unix Shell (.004), Python (.006), JavaScript (.007).", "mitigations": ["Code Signing", "Disable or Remove Feature", "Execution Prevention", "Antivirus/Antimalware"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1059.001": {"name": "PowerShell", "tactic": "Execution", "description": "Adversaries abuse PowerShell for execution and lateral movement. PowerShell is present on all modern Windows systems.", "mitigations": ["Code Signing", "Disable or Remove Feature", "Privileged Account Management", "Script Block Logging"], "platforms": ["Windows"]},
        "T1059.004": {"name": "Unix Shell", "tactic": "Execution", "description": "Adversaries abuse Unix shell commands and scripts for execution. Bash is the most common shell on Linux/macOS.", "mitigations": ["Execution Prevention", "Restrict File and Directory Permissions"], "platforms": ["Linux", "macOS"]},
        "T1059.007": {"name": "JavaScript", "tactic": "Execution", "description": "Adversaries abuse JavaScript for execution, often via browser-based attacks (XSS) or Node.js.", "mitigations": ["Disable or Remove Feature", "Restrict Web-Based Content"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1203": {"name": "Exploitation for Client Execution", "tactic": "Execution", "description": "Adversaries exploit vulnerabilities in client applications (browsers, PDF readers, Office) to execute arbitrary code.", "mitigations": ["Application Isolation and Sandboxing", "Exploit Protection"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1204": {"name": "User Execution", "tactic": "Execution", "description": "Adversaries rely on user interaction (opening files, clicking links) to execute malicious payloads.", "mitigations": ["Network Intrusion Prevention", "Restrict Web-Based Content", "User Training"], "platforms": ["Linux", "Windows", "macOS"]},
        # Persistence
        "T1053": {"name": "Scheduled Task/Job", "tactic": "Persistence", "description": "Adversaries abuse task scheduling (cron, Windows Task Scheduler, at) to execute programs at system startup or on a schedule.", "mitigations": ["Audit", "Operating System Configuration", "Privileged Account Management", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1136": {"name": "Create Account", "tactic": "Persistence", "description": "Adversaries create accounts to maintain access. Sub-techniques: Local Account (.001), Domain Account (.002), Cloud Account (.003).", "mitigations": ["MFA", "Network Segmentation", "Privileged Account Management"], "platforms": ["Linux", "Windows", "macOS", "Azure AD"]},
        "T1543": {"name": "Create or Modify System Process", "tactic": "Persistence", "description": "Adversaries create or modify system services (systemd, Windows Services, Launch Daemons) to persist.", "mitigations": ["Audit", "Restrict File and Directory Permissions", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1547": {"name": "Boot or Logon Autostart Execution", "tactic": "Persistence", "description": "Adversaries configure system settings to run programs at boot or logon. Sub-techniques: Registry Run Keys (.001), Login Items (.015).", "mitigations": ["Restrict Registry Permissions", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        # Privilege Escalation
        "T1068": {"name": "Exploitation for Privilege Escalation", "tactic": "Privilege Escalation", "description": "Adversaries exploit vulnerabilities in the OS kernel or privileged services to escalate privileges. Example: CVE-2024-1086 (Linux nf_tables UAF).", "mitigations": ["Application Isolation and Sandboxing", "Exploit Protection", "Update Software", "Threat Intelligence Program"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1548": {"name": "Abuse Elevation Control Mechanism", "tactic": "Privilege Escalation", "description": "Adversaries bypass elevation controls (UAC, sudo, setuid). Sub-techniques: Setuid/Setgid (.001), Bypass UAC (.002), Sudo Caching (.003).", "mitigations": ["Audit", "Execution Prevention", "Operating System Configuration", "Privileged Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1055": {"name": "Process Injection", "tactic": "Privilege Escalation", "description": "Adversaries inject code into running processes to escalate privileges or evade detection. Sub-techniques: DLL Injection (.001), Process Hollowing (.012), ptrace (.008).", "mitigations": ["Behavior Prevention on Endpoint", "Privileged Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        # Defense Evasion
        "T1070": {"name": "Indicator Removal", "tactic": "Defense Evasion", "description": "Adversaries delete or modify logs, artifacts, and indicators to cover their tracks. Sub-techniques: Clear Event Logs (.001), Clear Linux Logs (.002), File Deletion (.004).", "mitigations": ["Encrypt Sensitive Information", "Remote Data Storage", "Restrict File and Directory Permissions"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1027": {"name": "Obfuscated Files or Information", "tactic": "Defense Evasion", "description": "Adversaries obfuscate code, data, or commands to make detection and analysis difficult.", "mitigations": ["Antivirus/Antimalware", "Behavior Prevention on Endpoint"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1562": {"name": "Impair Defenses", "tactic": "Defense Evasion", "description": "Adversaries disable or modify security tools. Sub-techniques: Disable or Modify Tools (.001), Disable Windows Event Logging (.002), Disable Cloud Logs (.008).", "mitigations": ["Restrict File and Directory Permissions", "Restrict Registry Permissions", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        # Credential Access
        "T1110": {"name": "Brute Force", "tactic": "Credential Access", "description": "Adversaries use brute force to guess passwords. Sub-techniques: Password Guessing (.001), Password Cracking (.002), Password Spraying (.003), Credential Stuffing (.004).", "mitigations": ["Account Lockout Policies", "MFA", "Password Policies", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1555": {"name": "Credentials from Password Stores", "tactic": "Credential Access", "description": "Adversaries extract credentials from password stores (browsers, keychains, vaults). Sub-techniques: Keychain (.001), Web Browsers (.003), Password Managers (.005).", "mitigations": ["Password Policies", "Software Configuration"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1539": {"name": "Steal Web Session Cookie", "tactic": "Credential Access", "description": "Adversaries steal web session cookies to authenticate as a user without credentials.", "mitigations": ["Software Configuration", "User Training"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1003": {"name": "OS Credential Dumping", "tactic": "Credential Access", "description": "Adversaries dump credentials from the OS. Sub-techniques: LSASS Memory (.001), SAM (.002), /etc/passwd and /etc/shadow (.008).", "mitigations": ["Active Directory Configuration", "Credential Access Protection", "Operating System Configuration", "Privileged Account Management"], "platforms": ["Linux", "Windows"]},
        # Discovery
        "T1046": {"name": "Network Service Discovery", "tactic": "Discovery", "description": "Adversaries scan for services running on remote hosts using tools like nmap, netstat, or PowerShell.", "mitigations": ["Disable or Remove Feature", "Network Intrusion Prevention", "Network Segmentation"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1082": {"name": "System Information Discovery", "tactic": "Discovery", "description": "Adversaries gather system information (OS version, hardware, patches) to inform follow-on actions.", "mitigations": [], "platforms": ["Linux", "Windows", "macOS"]},
        "T1087": {"name": "Account Discovery", "tactic": "Discovery", "description": "Adversaries enumerate accounts to understand users and permissions. Sub-techniques: Local Account (.001), Domain Account (.002), Cloud Account (.004).", "mitigations": ["Operating System Configuration"], "platforms": ["Linux", "Windows", "macOS"]},
        # Lateral Movement
        "T1021": {"name": "Remote Services", "tactic": "Lateral Movement", "description": "Adversaries use valid accounts with remote services (SSH, RDP, SMB, WinRM) to move laterally. Sub-techniques: RDP (.001), SSH (.004), SMB (.002).", "mitigations": ["Disable or Remove Feature", "Limit Access to Resource Over Network", "MFA", "Network Segmentation", "Privileged Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1210": {"name": "Exploitation of Remote Services", "tactic": "Lateral Movement", "description": "Adversaries exploit vulnerabilities in remote services (SMB, SSH, databases) to move laterally.", "mitigations": ["Application Isolation and Sandboxing", "Disable or Remove Feature", "Exploit Protection", "Network Segmentation", "Privileged Account Management", "Update Software", "Vulnerability Scanning"], "platforms": ["Linux", "Windows", "macOS"]},
        # Collection
        "T1005": {"name": "Data from Local System", "tactic": "Collection", "description": "Adversaries search local system sources (files, databases, clipboards) for data of interest.", "mitigations": ["Data Loss Prevention"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1119": {"name": "Automated Collection", "tactic": "Collection", "description": "Adversaries use automated tools to collect data from the victim's environment at scale.", "mitigations": ["Data Loss Prevention", "Encrypt Sensitive Information", "Remote Data Storage"], "platforms": ["Linux", "Windows", "macOS"]},
        # Exfiltration
        "T1041": {"name": "Exfiltration Over C2 Channel", "tactic": "Exfiltration", "description": "Adversaries steal data by sending it over the existing command and control channel.", "mitigations": ["Data Loss Prevention", "Network Intrusion Prevention"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1567": {"name": "Exfiltration Over Web Service", "tactic": "Exfiltration", "description": "Adversaries exfiltrate data to legitimate cloud services (Google Drive, Dropbox, Pastebin). Sub-techniques: Exfiltration to Cloud Storage (.002).", "mitigations": ["Data Loss Prevention", "Restrict Web-Based Content"], "platforms": ["Linux", "Windows", "macOS"]},
        # Impact
        "T1486": {"name": "Data Encrypted for Impact", "tactic": "Impact", "description": "Adversaries encrypt data on target systems to disrupt operations (ransomware). Example: LockBit, BlackCat.", "mitigations": ["Backup", "Behavior Prevention on Endpoint", "Data Loss Prevention"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1489": {"name": "Service Stop", "tactic": "Impact", "description": "Adversaries stop services (databases, web servers, security tools) to disrupt operations or as a precursor to ransomware.", "mitigations": ["Network Segmentation", "Restrict File and Directory Permissions", "User Account Management"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1490": {"name": "Inhibit System Recovery", "tactic": "Impact", "description": "Adversaries delete or disable recovery mechanisms (Volume Shadow Copies, backups) to prevent system restoration.", "mitigations": ["Backup", "Operating System Configuration"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1499": {"name": "Endpoint Denial of Service", "tactic": "Impact", "description": "Adversaries perform denial of service to degrade or block availability. Sub-techniques: OS Exhaustion Flood (.001), Service Exhaustion Flood (.002), Application Exhaustion Flood (.003).", "mitigations": ["Filter Network Traffic", "Network Intrusion Prevention"], "platforms": ["Linux", "Windows", "macOS"]},
        "T1499.001": {"name": "Endpoint Denial of Service: OS Exhaustion Flood", "tactic": "Impact", "description": "Adversaries flood a service with requests that exhaust OS resources. Example: HTTP/2 Rapid Reset (CVE-2023-44487).", "mitigations": ["Rate limiting", "Connection limits", "WAF rules", "Upgrade protocol implementations"], "platforms": ["Linux", "Windows", "macOS"]},
    },
    "mitre_tactics": {
        "Initial Access": "Techniques adversaries use to gain an initial foothold in a network. Vectors include targeted spearphishing, exploiting public-facing applications, and supply chain compromise.",
        "Execution": "Techniques that result in adversary-controlled code running on a local or remote system. Often paired with Initial Access or Lateral Movement.",
        "Persistence": "Techniques adversaries use to keep access across system restarts, changed credentials, or other disruptions.",
        "Privilege Escalation": "Techniques adversaries use to gain higher-level permissions, often overlapping with Persistence.",
        "Defense Evasion": "Techniques adversaries use to avoid detection, including disabling security tools, obfuscating code, and clearing logs.",
        "Credential Access": "Techniques for stealing credentials (passwords, tokens, tickets) to access systems.",
        "Discovery": "Techniques adversaries use to gain knowledge about the system and network environment.",
        "Lateral Movement": "Techniques for moving through the network using legitimate credentials and remote services.",
        "Collection": "Techniques for gathering data of interest prior to exfiltration.",
        "Command and Control": "Techniques for communicating with compromised systems to maintain control.",
        "Exfiltration": "Techniques for stealing data from the target environment.",
        "Impact": "Techniques for disrupting, destroying, or manipulating systems and data (ransomware, DoS, data destruction).",
    },
    "platform_scanners": {
        "linux": {"vuln": ["Trivy", "Grype"], "sbom": ["Syft"], "config": ["Lynis", "kube-bench"], "network": ["nmap", "ss"]},
        "windows": {"vuln": ["Trivy", "Windows Update API"], "sbom": ["Syft"], "config": ["Windows Security Baselines"], "network": ["nmap", "Get-NetTCPConnection"]},
        "macos": {"vuln": ["Trivy", "Grype"], "sbom": ["Syft"], "config": ["Lynis"], "network": ["nmap"]},
        "docker": {"vuln": ["Trivy"], "sbom": ["Syft"], "config": ["Docker Bench"], "container": ["docker inspect"]},
        "k8s": {"vuln": ["Trivy"], "config": ["kube-bench", "Polaris"], "network": ["Calico audit"]},
    },
    "risk_thresholds": {"p0": 85, "p1": 70, "p2": 50, "p3": 0},
    "frameworks": ["NIST SP 800-53", "ISO/IEC 27001:2022", "MITRE ATT&CK", "CIS Benchmarks"],
}


def _ai_analyze_findings(findings: list[dict]) -> str:
    """Analyze a list of findings and produce a human-readable summary."""
    if not findings:
        return "No findings to analyze. The scan returned zero vulnerability results."
    critical = [f for f in findings if float(f.get("cvss_v3", 0)) >= 9.0]
    high = [f for f in findings if 7.0 <= float(f.get("cvss_v3", 0)) < 9.0]
    kev = [f for f in findings if f.get("kev")]
    lines = [f"**Analysis of {len(findings)} finding(s):**"]
    if critical:
        lines.append(f"🔴 **{len(critical)} CRITICAL** — These require immediate action (P0). "
                     f"CVEs: {', '.join(f.get('cve','?') for f in critical[:5])}")
    if high:
        lines.append(f"🟠 **{len(high)} HIGH** — Remediate within 7 days (P1). "
                     f"CVEs: {', '.join(f.get('cve','?') for f in high[:5])}")
    if kev:
        lines.append(f"⚠️ **{len(kev)} KEV-listed** (actively exploited in the wild) — "
                     f"These are being used by threat actors RIGHT NOW.")
    for f in findings[:3]:
        cve = f.get("cve", "")
        info = _AI_KNOWLEDGE["cve_db"].get(cve, {})
        if info:
            lines.append(f"\n**{cve}** — {info['title']}\n  Remediation: {info['remediation']}\n  MITRE: {info['mitre']}")
    return "\n".join(lines)


def _ai_platform_guidance(platform: str, actual_os: str) -> str:
    """Provide platform-specific guidance."""
    scanners = _AI_KNOWLEDGE["platform_scanners"].get(platform, {})
    lines = [f"**Platform: {platform.upper()}**"]
    if platform != actual_os and platform not in ("docker", "k8s", "cicd"):
        lines.append(f"⚠️ You're on **{actual_os}** but selected **{platform}**. "
                     f"For accurate results, deploy the agent on the actual target system "
                     f"or use Remote Deploy to generate commands for {platform}.")
    for collector, tools in scanners.items():
        lines.append(f"  • **{collector}**: {', '.join(tools)}")
    return "\n".join(lines)


def _ai_explain_score(score: float) -> str:
    """Explain a risk score in human terms."""
    if score >= 85:
        return (f"🔴 **Risk Score: {score:.1f}/100 (CRITICAL)**\n"
                f"This score indicates severe, exploitable vulnerabilities requiring immediate executive attention. "
                f"Your system is at high risk of compromise. Recommend: emergency patch cycle, incident readiness review.")
    if score >= 70:
        return (f"🟠 **Risk Score: {score:.1f}/100 (HIGH)**\n"
                f"Significant vulnerabilities exist. Remediation should begin within 7 days. "
                f"Focus on KEV-listed and internet-facing findings first.")
    if score >= 50:
        return (f"🟡 **Risk Score: {score:.1f}/100 (MEDIUM)**\n"
                f"Moderate risk. Schedule remediation within 30 days. "
                f"Prioritize findings with high EPSS scores (likelihood of exploitation).")
    return (f"🟢 **Risk Score: {score:.1f}/100 (LOW)**\n"
            f"Low risk profile. Continue monitoring. Address findings during regular maintenance windows.")


def _ai_process_message(message: str, context: dict) -> str:
    """Process a user message and return an AI response."""
    msg = message.lower().strip()
    page = context.get("page", "")
    findings = context.get("findings", [])
    platform = context.get("platform", "")
    actual_os = context.get("actual_os", "unknown")
    risk_score = context.get("risk_score")

    # Greeting
    if any(w in msg for w in ("hello", "hi ", "hey", "help", "what can you do")):
        return ("👋 **Hello! I'm the defriends AI Assistant.**\n\n"
                "I can help you with:\n"
                "• **Analyze findings** — Explain CVEs, risk scores, and remediation steps\n"
                "• **Deploy guidance** — Platform-specific agent deployment instructions\n"
                "• **Explain reports** — Break down assessment results in plain language\n"
                "• **Troubleshoot** — Diagnose connection errors, scanner issues, pipeline failures\n"
                "• **Compliance** — NIST SP 800-53 and ISO 27001 control mapping\n"
                "• **MITRE ATT&CK** — Explain techniques and attack patterns\n\n"
                "Try: *\"analyze my findings\"*, *\"explain the risk score\"*, "
                "*\"how do I deploy on Linux?\"*, or *\"what is T1195.002?\"*")

    # Risk score explanation
    if any(w in msg for w in ("risk score", "score", "risk level", "how bad")):
        if risk_score is not None:
            return _ai_explain_score(float(risk_score))
        return ("I don't have a risk score in context yet. Run the **Scoring** step first, "
                "or tell me a score: *\"explain score 82\"*")

    # Analyze findings
    if any(w in msg for w in ("analyze", "findings", "vulnerabilities", "what did", "scan results")):
        if findings:
            return _ai_analyze_findings(findings)
        return ("No findings in the current context. Please run the agent or ingestion step first. "
                "Go to **Deploy Agent** → Run Agent, or **Ingestion** → Submit Event.")

    # CVE lookup
    for cve_id, info in _AI_KNOWLEDGE["cve_db"].items():
        if cve_id.lower() in msg:
            return (f"**{cve_id}: {info['title']}**\n"
                    f"Severity: {info['severity'].upper()} (CVSS {info['cvss']})\n"
                    f"MITRE: {info['mitre']}\n"
                    f"Remediation: {info['remediation']}")

    # MITRE technique lookup — comprehensive database
    import re as _re
    technique_match = _re.search(r'(T\d{4}(?:\.\d{3})?)', msg.upper())
    if technique_match or any(w in msg for w in ("mitre", "att&ck", "attack", "tactic", "technique")):
        tech_db = _AI_KNOWLEDGE.get("mitre_techniques", {})
        tactics_db = _AI_KNOWLEDGE.get("mitre_tactics", {})

        # Try exact technique ID
        if technique_match:
            tid = technique_match.group(1)
            tech = tech_db.get(tid)
            if tech:
                lines = [
                    f"**{tid} — {tech['name']}**",
                    f"**Tactic:** {tech['tactic']}",
                    f"**Platforms:** {', '.join(tech.get('platforms', []))}",
                    f"\n{tech['description']}",
                ]
                if tech.get("mitigations"):
                    lines.append(f"\n**Mitigations:**")
                    for m in tech["mitigations"]:
                        lines.append(f"  • {m}")
                return "\n".join(lines)
            else:
                return f"I don't have details for technique **{tid}** in my database. Check the [MITRE ATT&CK site](https://attack.mitre.org/techniques/{tid.replace('.','/')})."

        # Tactic lookup
        for tactic_name, tactic_desc in tactics_db.items():
            if tactic_name.lower() in msg:
                lines = [f"**MITRE ATT&CK Tactic: {tactic_name}**\n{tactic_desc}\n\n**Techniques in this tactic:**"]
                for tid, tech in tech_db.items():
                    if tech["tactic"] == tactic_name:
                        lines.append(f"  • **{tid}** — {tech['name']}")
                return "\n".join(lines)

        # General MITRE question
        tactic_list = "\n".join(f"  • **{t}**: {d[:80]}..." for t, d in tactics_db.items())
        return (f"🗺️ **MITRE ATT&CK Framework**\n\n"
                f"MITRE ATT&CK is a knowledge base of adversary tactics and techniques based on real-world observations.\n"
                f"I know **{len(tech_db)}** techniques across **{len(tactics_db)}** tactics.\n\n"
                f"**Tactics:**\n{tactic_list}\n\n"
                f"Ask me about a specific technique (e.g., *\"What is T1190?\"*) or tactic (e.g., *\"explain Initial Access\"*).")

    # Platform / deployment
    if any(w in msg for w in ("deploy", "install", "agent", "linux", "windows", "macos", "kubernetes", "docker", "container")):
        target = "linux"
        if "windows" in msg: target = "windows"
        elif "mac" in msg: target = "macos"
        elif "kube" in msg or "k8s" in msg: target = "k8s"
        elif "docker" in msg or "container" in msg: target = "docker"
        guidance = _ai_platform_guidance(target, actual_os)
        bootstrap_base = "https://raw.githubusercontent.com/autobot786/secmesh_scaffold/master/secmesh_scaffold/agents/scripts"
        if target in ("linux", "macos"):
            guidance += (f"\n\n**Quick Install (one command):**\n"
                        f"```\ncurl -fsSL {bootstrap_base}/bootstrap.sh | bash -s -- "
                        f"--server YOUR_SERVER_URL --org YOUR_ORG --asset YOUR_ASSET --env prod\n```\n"
                        f"This auto-installs Python, Trivy, Syft, downloads the agent, and runs the scan.")
        elif target == "windows":
            guidance += (f"\n\n**Quick Install (PowerShell, one command):**\n"
                        f"```\nSet-ExecutionPolicy Bypass -Scope Process -Force; "
                        f"irm {bootstrap_base}/bootstrap.ps1 -OutFile $env:TEMP\\db.ps1; "
                        f"& $env:TEMP\\db.ps1 -Server YOUR_SERVER_URL -Org YOUR_ORG -Asset YOUR_ASSET -Env prod\n```\n"
                        f"This auto-installs dependencies and runs the scan.")
        return guidance

    # Troubleshooting
    if any(w in msg for w in ("error", "fail", "broken", "not working", "refused", "timeout", "can't connect")):
        return ("🔧 **Troubleshooting Guide:**\n\n"
                "**Connection refused / can't connect:**\n"
                "  • Verify the server is running: `python -m secmesh_scaffold`\n"
                "  • Check the URL: default is `http://127.0.0.1:8080`\n"
                "  • Windows: try `http://127.0.0.1:8080` instead of `localhost`\n\n"
                "**HTTP 422 (Validation Error):**\n"
                "  • Check JSON payload matches the EvidenceEvent schema\n"
                "  • Required fields: event_id, observed_at, asset.org_id, asset.asset_id, "
                "asset.environment, event_type\n"
                "  • source must be one of: sdk, agent, cicd, manual\n\n"
                "**Scanner not found:**\n"
                "  • Install Trivy: `curl -sfL https://...trivy.../install.sh | sh`\n"
                "  • Install Syft: `curl -sSfL https://...syft.../install.sh | sh`\n\n"
                "Tell me the specific error message and I can help further.")

    # Compliance
    if any(w in msg for w in ("nist", "iso 27001", "compliance", "framework", "gap analysis", "controls")):
        return ("📋 **Compliance Frameworks in defriends:**\n\n"
                "defriends auto-evaluates against two frameworks:\n\n"
                "**NIST SP 800-53:** SI-2 (Flaw Remediation), RA-5 (Vulnerability Scanning), "
                "SC-7 (Boundary Protection), CM-3 (Configuration Management), CA-7 (Continuous Monitoring)\n\n"
                "**ISO/IEC 27001:2022 Annex A:** A.8.8 (Technical Vulnerabilities), A.5.7 (Threat Intelligence), "
                "A.8.20 (Network Security), A.8.9 (Configuration Management), A.8.16 (Monitoring)\n\n"
                "The gap analysis report includes a **NIST ↔ ISO cross-reference table** and "
                "per-control pass/partial/fail status with remediation guidance.")

    # What is defriends
    if any(w in msg for w in ("what is dirtybots", "about", "what does this do", "how does it work")):
        return ("🛡️ **defriends** is a security assessment platform that:\n\n"
                "1. **Ingests** security evidence (vulnerability scans, SBOMs, config audits)\n"
                "2. **Normalizes** the data into a standard format\n"
                "3. **Maps** findings to MITRE ATT&CK techniques\n"
                "4. **Scores** risk using CVSS + EPSS + KEV + exposure context\n"
                "5. **Reports** with PDF generation, dual-framework gap analysis (NIST + ISO 27001), "
                "and prioritized remediation recommendations\n\n"
                "The agent can be deployed on Linux, Windows, macOS, containers, Kubernetes, and CI/CD pipelines.")

    # Explain score number
    import re
    score_match = re.search(r'(\d+\.?\d*)\s*/?\s*100|score\s*(\d+\.?\d*)', msg)
    if score_match:
        s = float(score_match.group(1) or score_match.group(2))
        return _ai_explain_score(s)

    # Default
    return ("I'm not sure I understood that. I can help with:\n"
            "• **\"analyze findings\"** — Review current scan results\n"
            "• **\"explain risk score\"** — Interpret the risk rating\n"
            "• **\"CVE-2024-3094\"** — Look up a specific CVE\n"
            "• **\"deploy on linux\"** — Platform deployment guidance\n"
            "• **\"troubleshoot errors\"** — Diagnose issues\n"
            "• **\"NIST compliance\"** — Framework gap analysis info\n"
            "• **\"what is T1195.002\"** — MITRE technique details")


@app.post("/v1/ai/assist", tags=["ai"])
def ai_assist(body: dict):
    """AI Assistant endpoint — processes natural language queries."""
    message = body.get("message", "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")
    context = body.get("context", {})
    try:
        from services.ai_assistant.app.assistant import answer as _structured_answer
        structured = _structured_answer(message, context)
        response = structured.get("reply", "")
    except Exception:
        structured = {}
        response = _ai_process_message(message, context)

    # Store in FAQ
    _faq_store.append({
        "id": f"faq-{uuid.uuid4().hex[:8]}",
        "question": message,
        "answer": response,
        "asked_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "portal": context.get("portal", "unknown"),
    })

    return {
        "response": response,
        "source": "dirtybots-ai",
        "actions": structured.get("actions", []),
        "suggestions": structured.get("suggestions", []),
        "sanitization": structured.get("sanitization", {"prompt_injection_flags": []}),
    }


# ---------------------------------------------------------------------------
# FAQ — Frequently Asked Questions (auto-populated from AI Assistant)
# ---------------------------------------------------------------------------
_faq_store: list[dict] = []


@app.get("/v1/faq", tags=["faq"])
def list_faq(limit: int = 50):
    """Return all FAQ entries, most recent first. Duplicate questions are merged."""
    seen: dict[str, dict] = {}
    for entry in _faq_store:
        key = entry["question"].strip().lower()
        if key not in seen:
            seen[key] = {**entry, "count": 1}
        else:
            seen[key]["count"] += 1
    items = sorted(seen.values(), key=lambda x: x.get("count", 0), reverse=True)
    return {"faq": items[:limit], "total": len(items)}


@app.delete("/v1/faq/{faq_id}", tags=["faq"])
def delete_faq(faq_id: str):
    """Delete a single FAQ entry (admin use)."""
    global _faq_store
    before = len(_faq_store)
    _faq_store = [f for f in _faq_store if f["id"] != faq_id]
    if len(_faq_store) == before:
        raise HTTPException(status_code=404, detail="FAQ entry not found")
    return {"deleted": True, "faq_id": faq_id}
